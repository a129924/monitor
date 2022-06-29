# %%
from itsdangerous import want_bytes
import openpyxl
import configparser, urllib.parse
from pyodbc import DatabaseError

from sqlalchemy import create_engine
# %%
class read_excel():
    def __init__(self,filename, sheetname=0):
        self._filename = filename
        self._sheetname = sheetname
        
    @property
    def filename(self):
        return self._filename
    @property
    def sheetname(self):
        return self._sheetname
    
    @property
    def rows(self):
        return self._sheet.max_row
    
    @property
    def columns(self):
        return self._sheet.max_column
    
    def open(self):
        self._workbook = openpyxl.load_workbook(self._filename)
        self._sheet = self._workbook.worksheets[self._sheetname]

    def read(self):
        self.open()
        print(self.rows,self.columns)
        datas = []
        for row in range(1,self.rows+1,1):
            new_data = [[]] * self.columns
            for col in range(1,self.columns+1,1):
                new_data[col-1].append(self._sheet.cell(row=row, column=col).value)
                
            datas.append(new_data)
            
        return datas
# %%
if "__main__" == __name__:
    
    filepath = r".\file\20220620.xlsx"
    data  = read_excel(filepath)
    data = data.read()
    
# %%



# parser = configparser.ConfigParser()
# parser.read(r".\dbSetting\db.ini")

# mssql_conf = parser["sqlserver"]

# params = urllib.parse.quote_plus(
#     f"DRIVER={mssql_conf['driver']};SERVER={mssql_conf['server']};DATABASE={mssql_conf['database']};UID={mssql_conf['uid']};PWD={mssql_conf['pwd']}")
# engine = create_engine("mssql+pyodbc:///?odbc_connect=%s" % params)

# %%
